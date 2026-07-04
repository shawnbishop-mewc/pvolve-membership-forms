# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.utils import get_column_letter

INK="FF1C1B19"; CLAY="FFB15C3E"; SAND="FFF3EEE7"; SANDD="FFEAE3D8"
GOOD="FFDCE6D0"; GOODT="FF3C5A2E"; BAD="FFF1D6CD"; BADT="FF8A3A22"; MID="FFEDE7DB"
WHITE="FFFBF9F5"; LINE="FFDAD2C4"

wb=Workbook()

# ---------------- Submissions ----------------
ws=wb.active; ws.title="Submissions"
headers=["Submitted","Request Type","Converted From Cancel?","Location","Member Name",
 "Email","Phone","Freeze Start","Freeze Duration","Reasons (all checked)",
 "Rating: Quality of Workout","Rating: Front Desk Service","Rating: Class Times",
 "Rating: Cleanliness","Rating: Overall Experience","Return Likelihood (1-10)",
 "Comments","Signature","Acknowledged"]
ws.append(headers)

hdr_fill=PatternFill("solid",fgColor=INK)
hdr_font=Font(name="Calibri",bold=True,color=WHITE,size=11)
for c in range(1,len(headers)+1):
    cell=ws.cell(row=1,column=c)
    cell.fill=hdr_fill; cell.font=hdr_font
    cell.alignment=Alignment(horizontal="left",vertical="center",wrap_text=True)
ws.row_dimensions[1].height=34

rows=[
 ["2026-07-03 14:22","Freeze","No","Post Oak","Jenna Lowe","jenna.lowe@email.com","(713) 555-0142",
  "2026-07-20","30 days","Travel; No Time to Attend",
  "Great","Great","Good","Great","Great",9,
  "Just traveling for work through the summer — back in August!","Jenna Lowe","Yes"],
 ["2026-07-03 14:41","Cancellation","No","Memorial","Sarah Mitchell","sarah.mitchell@email.com","(713) 555-1234",
  "","","No Time to Attend; Financial; Convenient Classes Unavailable",
  "Great","Good","Poor","Fair","Fair",7,
  "Love the workouts but the only slots that fit me are always full. More evening classes and I'd stay.","Sarah Mitchell","Yes"],
 ["2026-07-03 15:08","Freeze","Yes","Post Oak","Maria Gomez","maria.gomez@email.com","(713) 555-0199",
  "2026-08-01","60 days","Financial",
  "Good","Good","Good","Good","Good",6,
  "Money is tight right now but I really don't want to lose my rate.","Maria Gomez","Yes"],
 ["2026-07-02 09:15","Cancellation","No","Post Oak","Beth Alvarez","beth.a@email.com","(713) 555-0177",
  "","","Moving or Relocating",
  "Great","Great","Great","Great","Great",4,
  "Moving to Dallas. Wish you had a studio there — best workout I've found.","Beth Alvarez","Yes"],
]
for r in rows: ws.append(r)

# widths
widths=[16,15,12,11,17,26,15,13,13,34,13,13,12,12,13,12,42,16,12]
for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w

# body style
thin=Side(style="thin",color=LINE)
for row in ws.iter_rows(min_row=2,max_row=ws.max_row,max_col=len(headers)):
    for cell in row:
        cell.alignment=Alignment(vertical="top",wrap_text=True)
        cell.border=Border(bottom=thin)
        cell.font=Font(size=10)

# conditional formatting on rating columns K:O (11-15)
for col in range(11,16):
    L=get_column_letter(col); rng=f"{L}2:{L}1000"
    ws.conditional_formatting.add(rng,CellIsRule(operator="equal",formula=['"Great"'],fill=PatternFill("solid",fgColor=GOOD),font=Font(color=GOODT)))
    ws.conditional_formatting.add(rng,CellIsRule(operator="equal",formula=['"Good"'],fill=PatternFill("solid",fgColor=GOOD),font=Font(color=GOODT)))
    ws.conditional_formatting.add(rng,CellIsRule(operator="equal",formula=['"Fair"'],fill=PatternFill("solid",fgColor=MID)))
    ws.conditional_formatting.add(rng,CellIsRule(operator="equal",formula=['"Poor"'],fill=PatternFill("solid",fgColor=BAD),font=Font(color=BADT)))
    ws.conditional_formatting.add(rng,CellIsRule(operator="equal",formula=['"Unsatisfactory"'],fill=PatternFill("solid",fgColor=BAD),font=Font(bold=True,color=BADT)))
# likelihood color scale P (16)
ws.conditional_formatting.add("P2:P1000",ColorScaleRule(
    start_type="num",start_value=1,start_color="FFE8B4A2",
    mid_type="num",mid_value=5,mid_color="FFF3EAD9",
    end_type="num",end_value=10,end_color="FFCFE0C2"))

ws.freeze_panes="A2"
ws.auto_filter.ref=f"A1:{get_column_letter(len(headers))}1"

# ---------------- Dashboard ----------------
db=wb.create_sheet("Dashboard")
db.column_dimensions["A"].width=34
db.column_dimensions["B"].width=14
db.column_dimensions["C"].width=14
db.column_dimensions["D"].width=40

def title(cell,text):
    db[cell]=text; db[cell].font=Font(bold=True,size=16,color=INK)
def sub(cell,text):
    db[cell]=text; db[cell].font=Font(bold=True,size=11,color=WHITE); db[cell].fill=PatternFill("solid",fgColor=CLAY)
    db[cell].alignment=Alignment(horizontal="left",vertical="center")
def lbl(cell,text):
    db[cell]=text; db[cell].font=Font(size=11,color="FF1C1B19")
def val(cell,formula,bold=True):
    db[cell]=formula; db[cell].font=Font(size=12,bold=bold,color=INK)

S="Submissions"
title("A1","Pvolve · Member Retention Dashboard")
db["A2"]="Live figures — update automatically as rows are added to the Submissions tab."
db["A2"].font=Font(italic=True,size=10,color="FF5C574F")

sub("A4","VOLUME"); db.merge_cells("A4:D4")
lbl("A5","Total requests"); val("B5",f'=COUNTA({S}!A2:A100000)')
lbl("A6","Freeze requests"); val("B6",f'=COUNTIF({S}!B2:B100000,"Freeze")')
lbl("A7","Cancellation requests"); val("B7",f'=COUNTIF({S}!B2:B100000,"Cancellation")')
lbl("A8","Saves (cancel → froze instead)"); val("B8",f'=COUNTIF({S}!C2:C100000,"Yes")')
lbl("A9","Save rate on cancel attempts"); val("B9",f'=IFERROR(B8/(B7+B8),0)')
db["B9"].number_format="0%"

sub("A11","BY LOCATION"); db.merge_cells("A11:D11")
lbl("A12","Memorial — total"); val("B12",f'=COUNTIF({S}!D2:D100000,"Memorial")')
lbl("A13","Memorial — cancellations"); val("B13",f'=COUNTIFS({S}!D2:D100000,"Memorial",{S}!B2:B100000,"Cancellation")')
lbl("A14","Post Oak — total"); val("B14",f'=COUNTIF({S}!D2:D100000,"Post Oak")')
lbl("A15","Post Oak — cancellations"); val("B15",f'=COUNTIFS({S}!D2:D100000,"Post Oak",{S}!B2:B100000,"Cancellation")')

sub("A17","TOP REASONS  (count of times checked)"); db.merge_cells("A17:D17")
reasons=["Moving or Relocating","No Time to Attend","Financial","Medical","Seasonal Resident",
 "Fitness Routine Change","Convenient Classes Unavailable","Unsatisfied with Facility",
 "Unsatisfied with Studio Team","Travel","Upgrade/Downgrade","Other"]
r=18
for reason in reasons:
    lbl(f"A{r}",reason)
    val(f"B{r}",f'=SUMPRODUCT(--ISNUMBER(SEARCH("{reason}",{S}!$J$2:$J$100000)))',bold=False)
    r+=1

sub("A31","SERVICE RATINGS  (# rating Poor or Unsatisfactory = trouble spots)"); db.merge_cells("A31:D31")
cats=[("Quality of Workout","K"),("Front Desk Service","L"),("Class Times","M"),
      ("Cleanliness","N"),("Overall Experience","O")]
r=32
for name,col in cats:
    lbl(f"A{r}",name)
    val(f"B{r}",f'=COUNTIF({S}!{col}2:{col}100000,"Poor")+COUNTIF({S}!{col}2:{col}100000,"Unsatisfactory")',bold=False)
    db[f"C{r}"]="Poor/Unsat."; db[f"C{r}"].font=Font(size=9,italic=True,color="FF5C574F")
    r+=1

sub("A38","RETURN LIKELIHOOD"); db.merge_cells("A38:D38")
lbl("A39","Average — all requests"); val("B39",f'=IFERROR(ROUND(AVERAGE({S}!P2:P100000),1),0)')
lbl("A40","Average — cancellations only")
val("B40",f'=IFERROR(ROUND(AVERAGEIF({S}!B2:B100000,"Cancellation",{S}!P2:P100000),1),0)')

# ---------------- Legend ----------------
lg=wb.create_sheet("Legend")
lg.column_dimensions["A"].width=30; lg.column_dimensions["B"].width=80
lg["A1"]="How this workbook is used"; lg["A1"].font=Font(bold=True,size=15)
notes=[
 ("Submissions tab","One row per submitted form. This is where the website automation writes each new freeze or cancellation. Do not rename the column headers — the automation maps to them."),
 ("Request Type","'Freeze' or 'Cancellation'. A member who started a cancellation and chose to freeze instead is saved as 'Freeze' with 'Converted From Cancel?' = Yes."),
 ("Converted From Cancel?","'Yes' means this was a rescued cancellation — the freeze-instead nudge worked."),
 ("Reasons (all checked)","All reasons the member selected, separated by '; '. The Dashboard counts these with a text search, so keep them in one cell."),
 ("Ratings K–O","Great / Good / Fair / Poor / Unsatisfactory. Auto-colored — green is good, red flags Poor/Unsatisfactory at a glance."),
 ("Return Likelihood","1–10 self-reported. Color scale: red = low, green = high."),
 ("Dashboard tab","Updates automatically. Watch the Save Rate, Top Reasons, and which rating category collects the most Poor/Unsatisfactory marks — that's your churn driver."),
 ("Uploading to Google Sheets","Open Google Drive → New → File upload → this .xlsx, then open with Google Sheets (or Sheets → File → Import). Formatting and formulas carry over."),
]
r=3
for k,v in notes:
    lg[f"A{r}"]=k; lg[f"A{r}"].font=Font(bold=True,size=11); lg[f"A{r}"].alignment=Alignment(vertical="top")
    lg[f"B{r}"]=v; lg[f"B{r}"].alignment=Alignment(wrap_text=True,vertical="top")
    lg.row_dimensions[r].height=46
    r+=1

out=r"C:\Users\ShawnBishop\Claude Code\pvolve-forms\Pvolve-Submissions-Sheet.xlsx"
wb.save(out)
print("Saved:",out)
